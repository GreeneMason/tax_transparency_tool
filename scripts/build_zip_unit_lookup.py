from __future__ import annotations

import argparse
import csv
import io
import zipfile
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

PID_WIDTH = 146


def digits_only(value: str) -> str:
    return "".join(char for char in value if char.isdigit())


def normalize_zip_code(value: str) -> str:
    numeric = digits_only(value)
    if len(numeric) < 5:
        return ""
    return numeric[:5]


def normalize_fips5(value: str) -> str:
    numeric = digits_only(value)
    if len(numeric) < 5:
        return ""
    return numeric[:5]


def parse_pid_line(line: str) -> dict[str, str]:
    record = line.rstrip("\r\n")
    if not record or len(record) < 12:
        return {}

    if len(record) < PID_WIDTH:
        record = record.ljust(PID_WIDTH)

    def normalized(segment: str) -> str:
        return " ".join(segment.strip().split())

    unit_id = record[0:12].strip()
    if len(unit_id) != 12:
        return {}

    return {
        "unit_id": unit_id,
        "unit_name": normalized(record[12:76]),
        "county_name": normalized(record[76:111]),
        "place_fips": record[111:116].strip(),
    }


def load_pid_units_by_county(pid_path: Path, state_fips_filter: str) -> dict[str, list[dict[str, str]]]:
    county_to_units: dict[str, list[dict[str, str]]] = {}
    with pid_path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            row = parse_pid_line(line)
            if not row:
                continue

            unit_id = row["unit_id"]
            state_fips = unit_id[0:2]
            county_fips = unit_id[3:6]

            if state_fips != state_fips_filter:
                continue

            county_full = f"{state_fips}{county_fips}"
            county_to_units.setdefault(county_full, []).append(
                {
                    "unit_id": unit_id,
                    "unit_name": row["unit_name"],
                    "county_name": row["county_name"],
                    "place_fips": row["place_fips"],
                }
            )
    return county_to_units


def open_hud_csv_reader(hud_path: Path) -> tuple[csv.DictReader, io.TextIOBase, str]:
    if hud_path.suffix.lower() == ".zip":
        archive = zipfile.ZipFile(hud_path)
        csv_members = [name for name in archive.namelist() if name.lower().endswith(".csv")]
        if not csv_members:
            archive.close()
            raise FileNotFoundError(f"No CSV found in archive: {hud_path}")

        member = csv_members[0]
        raw = archive.open(member)
        text = io.TextIOWrapper(raw, encoding="utf-8", errors="replace", newline="")
        reader = csv.DictReader(text)

        class ZipTextWrapper(io.TextIOBase):
            def close(self) -> None:
                try:
                    text.close()
                finally:
                    archive.close()

        return reader, ZipTextWrapper(), member

    text = hud_path.open("r", encoding="utf-8", errors="replace", newline="")
    reader = csv.DictReader(text)
    return reader, text, hud_path.name


def pick_column(fieldnames: list[str], candidates: list[str]) -> str:
    index = {name.upper().strip(): name for name in fieldnames}
    for candidate in candidates:
        if candidate in index:
            return index[candidate]
    return ""


def parse_hud_zip_county(
    hud_path: Path,
    state_fips_filter: str,
) -> tuple[dict[tuple[str, str], float], str]:
    zip_county_ratios: dict[tuple[str, str], float] = {}
    source_name = hud_path.name

    def ingest_rows(fieldnames: list[str], rows: list[dict[str, str]]) -> None:
        zip_col = pick_column(fieldnames, ["ZIP", "ZIPCODE", "ZIP_CODE", "ZCTA5", "ZCTA"])
        county_col = pick_column(fieldnames, ["COUNTY", "COUNTY_FIPS", "COUNTYFP", "CNTY"])
        ratio_col = pick_column(fieldnames, ["TOT_RATIO", "RES_RATIO", "RATIO"])

        if not zip_col or not county_col:
            raise ValueError(
                "Unable to find required ZIP/COUNTY columns in HUD file. "
                "Expected headers such as ZIP and COUNTY."
            )

        for row in rows:
            zip_code = normalize_zip_code(row.get(zip_col, ""))
            county_fips_full = normalize_fips5(row.get(county_col, ""))
            if not zip_code or not county_fips_full:
                continue

            if county_fips_full[0:2] != state_fips_filter:
                continue

            ratio = 1.0
            if ratio_col:
                raw_ratio = (row.get(ratio_col) or "").strip()
                if raw_ratio:
                    try:
                        ratio = float(raw_ratio)
                    except ValueError:
                        ratio = 0.0

            key = (zip_code, county_fips_full)
            previous = zip_county_ratios.get(key)
            if previous is None or ratio > previous:
                zip_county_ratios[key] = ratio

    if hud_path.suffix.lower() in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise ImportError(
                "Reading .xlsx HUD files requires openpyxl. Install with: pip install openpyxl"
            )

        workbook = load_workbook(hud_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            header_row = next(iterator, None)
            if header_row is None:
                raise ValueError(f"HUD file has no header row: {hud_path}")

            fieldnames = [str(value).strip() if value is not None else "" for value in header_row]
            rows: list[dict[str, str]] = []

            for values in iterator:
                row: dict[str, str] = {}
                for index, field in enumerate(fieldnames):
                    if not field:
                        continue
                    value = values[index] if index < len(values) else None
                    row[field] = "" if value is None else str(value).strip()
                rows.append(row)

            ingest_rows(fieldnames, rows)
            return zip_county_ratios, source_name
        finally:
            workbook.close()

    reader, closer, source_name = open_hud_csv_reader(hud_path)
    try:
        if not reader.fieldnames:
            raise ValueError(f"HUD file has no header row: {hud_path}")

        fieldnames = [name.strip() for name in reader.fieldnames]
        rows: list[dict[str, str]] = []
        for row in reader:
            normalized_row: dict[str, str] = {}
            for key, value in row.items():
                if not key:
                    continue
                normalized_row[key.strip()] = "" if value is None else str(value).strip()
            rows.append(normalized_row)

        ingest_rows(fieldnames, rows)
        return zip_county_ratios, source_name
    finally:
        closer.close()


def build_zip_unit_rows(
    zip_county_ratios: dict[tuple[str, str], float],
    county_to_units: dict[str, list[dict[str, str]]],
    state_fips_filter: str,
    source_name: str,
) -> list[dict[str, str]]:
    by_key: dict[tuple[str, str], dict[str, str]] = {}

    for (zip_code, county_fips_full), ratio in zip_county_ratios.items():
        county_units = county_to_units.get(county_fips_full, [])
        if not county_units:
            continue

        for unit in county_units:
            unit_id = unit["unit_id"]
            key = (zip_code, unit_id)
            record = {
                "zip_code": zip_code,
                "state_fips": state_fips_filter,
                "county_fips": county_fips_full[2:5],
                "county_fips_full": county_fips_full,
                "unit_id": unit_id,
                "hud_ratio": f"{ratio:.6f}",
                "match_scope": "county",
                "source_file": source_name,
            }

            previous = by_key.get(key)
            if previous is None or float(record["hud_ratio"]) > float(previous["hud_ratio"]):
                by_key[key] = record

    rows = list(by_key.values())
    rows.sort(key=lambda row: (row["zip_code"], -float(row["hud_ratio"]), row["unit_id"]))
    return rows


def write_rows(rows: list[dict[str, str]], output_path: Path) -> None:
    fieldnames = [
        "zip_code",
        "state_fips",
        "county_fips",
        "county_fips_full",
        "unit_id",
        "hud_ratio",
        "match_scope",
        "source_file",
    ]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build ZIP-to-unit lookup CSV from HUD ZIP-COUNTY crosswalk and Fin_PID_2023."
    )
    parser.add_argument(
        "--hud-zip-county",
        required=True,
        help="Path to HUD ZIP-COUNTY crosswalk CSV, ZIP archive, or XLSX file.",
    )
    parser.add_argument(
        "--pid",
        default="data/Fin_PID_2023.txt",
        help="Path to Fin_PID_2023 fixed-width file.",
    )
    parser.add_argument(
        "--state-fips",
        default="53",
        help="2-digit state FIPS code filter (default: 53 for Washington).",
    )
    parser.add_argument(
        "--output",
        default="data/output/wa_zip_to_unit_lookup.csv",
        help="Output CSV path for db/load_zip_lookup.sql.",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    hud_path = Path(args.hud_zip_county)
    pid_path = Path(args.pid)
    output_path = Path(args.output)
    state_fips_filter = args.state_fips.strip()

    if len(state_fips_filter) != 2 or not state_fips_filter.isdigit():
        raise ValueError("--state-fips must be a 2-digit numeric FIPS code.")

    if not hud_path.exists():
        raise FileNotFoundError(f"HUD ZIP-COUNTY file not found: {hud_path}")
    if not pid_path.exists():
        raise FileNotFoundError(f"PID file not found: {pid_path}")

    county_to_units = load_pid_units_by_county(pid_path, state_fips_filter)
    zip_county_ratios, source_name = parse_hud_zip_county(hud_path, state_fips_filter)
    rows = build_zip_unit_rows(zip_county_ratios, county_to_units, state_fips_filter, source_name)
    write_rows(rows, output_path)

    zip_count = len({row["zip_code"] for row in rows})
    unit_count = len({row["unit_id"] for row in rows})

    print(f"Built {len(rows):,} ZIP→unit rows for state {state_fips_filter}.")
    print(f"Distinct ZIPs: {zip_count:,} | Distinct units: {unit_count:,}")
    print(f"Output: {output_path}")
    print("Note: ZIP-COUNTY gives county-level linkage; place-level precision needs an additional ZIP→place/ZCTA-to-place dataset.")


if __name__ == "__main__":
    main()
